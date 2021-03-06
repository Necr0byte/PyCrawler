import gspread, os, openpyxl, subprocess
from openpyxl import *
from oauth2client.service_account import ServiceAccountCredentials

#variables globales:
stocks = ["AAPL", "ABEV", "ABT", "ADBE"]#, "AGRO", "ADP", "AIG", "AMD", "AMGN", "AMX", "AMZN", "ARCO", "AUY", "AXP", "AZN", "BA", "BABA", "BBD", "BBDD", "BBVA", "BCS", "BHP", "BIDU", "BIIB", "BNG", "BP", "BRFS", "BSBR", "C", "CAT", "CHL", "CL", "COST", "CRM", "CSCO", "CVX", "CX", "DESP", "DISN", "DISND", "EBAY", "ERJ", "FB", "FDX", "FMX", "GE", "GGB", "GILD", "GILDD", "GLNT", "GLNTD", "GOLD", "GOLDD", "GOOGL", "GS", "GSK", "HD", "HMC", "HMY", "HPQ", "HSBC", "IBM", "INFY", "INTC", "ITUB", "JD", "JNJ", "JPM", "KO", "KOD", "LMT", "LVS", "LYG", "MCD", "MDT", "MELI", "MELID", "MMM", "MO", "MRK", "MSFT", "MSFTD", "NEM", "NFLX", "NKE", "NOKA", "NVDA", "NVS", "OGZD", "ORCL", "PBR", "PEP", "PFE", "PG", "PYPL", "QCOM", "RDS", "RIO", "RTX", "SAN", "SAP", "SBS", "SBUX", "SID", "SLB", "SNAP", "SNE", "T", "TEN", "TGT", "TM", "TOT", "TRIP", "TSLA", "TSLAD", "TSM", "TSU", "TWTR", "TXN", "TXR", "UGP", "UN", "V", "VALE", "VIST", "VIV", "VOD", "VRSN", "VZ", "WFC", "WMT", "WMTD", "X", "XOM", "XROX", "YELP"]
refused_list = []
invalid_list = []
pending_list = stocks.copy()

#Variables para el google spreadsheets:
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('secret.json',scope)
client = gspread.authorize(creds)

#chunk of code to create an Excel file
database = openpyxl.Workbook()
database.create_sheet(index = 0 , title = "CEDEARs")
cedears = database.active
#database.save("C:\\Users\\Magnus\\VCS\\PyCrawler\\Stocks.xlsx")

def spreadsheets(task, col, row, data):
    rowcol=row+str(col+1)
    if task == 'put':
        print (rowcol, " = ", data)
        cedears[rowcol] = data
    elif task == 'get':
        return cedears[rowcol]
    else:
        print('tasks must be either "get" or "put"')
    database.save("C:\\Users\\Magnus\\VCS\\PyCrawler\\Stocks.xlsx")

#This function is deprecated: 
def google_spreadsheets(task, col, row, data):
    sheet = client.open("CEDEARs").get_worksheet(1)
    if task == 'put':
        sheet.update_cell(row, col, data)
    elif task == 'get':
        return sheet.cell(row, col)
    else:
        print('tasks must be either "get" or "put"')

def get_div(ticker):

    grabber = 'python historical_grabber.py' + " " + ticker
    #p = subprocess.Popen(grabber, stdout=subprocess.PIPE, shell=True)
    try:
        div = subprocess.check_output(grabber)
        div = div.strip()
        # div = str(div)
        # div = div.replace('\n','')
        # div = div.replace('\r','')
        # div = div.replace('b','')
        # div = div.replace('\b','')
        # div = div.replace('\'','')
        # div = div.replace(' ','')
        return str(div)
    except:
        return 0

while (pending_list):
    for i in pending_list:
        d = get_div(i)
        ix = stocks.index(i)
        spreadsheets('put', ix, 'A', i)
        spreadsheets('put', ix, 'B', d)
        pending_list.remove(i)